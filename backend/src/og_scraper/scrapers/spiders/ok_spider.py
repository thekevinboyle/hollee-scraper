"""Oklahoma Corporation Commission (OCC) spider for bulk data downloads.

Targets the OCC's public bulk data files including:
- RBDMS Well Data (CSV, nightly)
- Incident Reports (CSV, daily)
- Intent to Drill permits (XLSX, daily)
- Well Completions (XLSX, daily)
- Operator/Purchaser/Plugger lists (XLSX, daily)
- UIC injection volumes (XLSX)
- Well Transfers (XLSX, daily)
- Orphan wells, State Funds wells (XLSX, weekly)

Production data is maintained by the Oklahoma Tax Commission (OkTAP),
not the OCC, and is stubbed here for future Playwright-based access.
"""

import csv
import io
import logging
from datetime import date, datetime

import scrapy

try:
    import openpyxl
except ImportError:
    openpyxl = None

from og_scraper.scrapers.items import DocumentItem, WellItem
from og_scraper.scrapers.spiders.base import BaseOGSpider

logger = logging.getLogger(__name__)


class OklahomaOCCSpider(BaseOGSpider):
    """Spider for Oklahoma Corporation Commission bulk data downloads.

    Downloads CSV and XLSX files from static OCC URLs. No authentication
    or JavaScript rendering required -- pure HTTP downloads.

    Spider type: BulkDownloadSpider
    State FIPS code: 35
    """

    name = "ok_occ"
    state_code = "OK"
    state_name = "Oklahoma"
    agency_name = "Corporation Commission (OCC)"
    base_url = "https://oklahoma.gov/occ/divisions/oil-gas.html"
    requires_playwright = False

    custom_settings = {
        "DOWNLOAD_DELAY": 3,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "AUTOTHROTTLE_ENABLED": True,
        "AUTOTHROTTLE_START_DELAY": 3,
        "AUTOTHROTTLE_MAX_DELAY": 30,
        "AUTOTHROTTLE_TARGET_CONCURRENCY": 2.0,
    }

    BASE_URL = "https://oklahoma.gov"

    # Map of dataset name -> (relative_url, format, report_type)
    BULK_FILES = {
        # Well Information
        "rbdms_wells": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/rbdms-wells.csv",
            "csv",
            "well_data",
        ),
        "incidents": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/ogcd-incidents.csv",
            "csv",
            "incident_report",
        ),
        # Intent to Drill & Completions
        "itd_master": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/ITD-wells-formations-base.xlsx",
            "xlsx",
            "well_permit",
        ),
        "completions_master": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/completions-wells-formations-base.xlsx",
            "xlsx",
            "completion_report",
        ),
        "well_transfers": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/well-transfers-daily.xlsx",
            "xlsx",
            "well_transfer",
        ),
        # Operators
        "operators": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/operator-list.xlsx",
            "xlsx",
            "operator_list",
        ),
        # UIC
        "uic_wells": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/online-active-well-list.xlsx",
            "xlsx",
            "uic_data",
        ),
        "uic_2025": (
            "/content/dam/ok/en/occ/documents/og/ogdatafiles/2025-uic-injection-volumes.xlsx",
            "xlsx",
            "uic_injection",
        ),
    }

    # --- Production data stubs (Oklahoma Tax Commission, not OCC) ---
    OKTAP_URL = "https://oktap.tax.ok.gov/OkTAP/web?link=PUBLICPUNLKP"
    GROSS_PRODUCTION_URL = "https://otcportal.tax.ok.gov/gpx/index.php"

    def start_requests(self):
        """Yield one request per bulk file in BULK_FILES."""
        for dataset_name, (url_path, fmt, report_type) in self.BULK_FILES.items():
            yield scrapy.Request(
                url=f"{self.BASE_URL}{url_path}",
                callback=self.parse_bulk_file,
                meta={
                    "dataset_name": dataset_name,
                    "file_format": fmt,
                    "report_type": report_type,
                },
                errback=self.errback_handler,
            )

    def start_oktap_requests(self):
        """Stub: Production data from Oklahoma Tax Commission.

        OkTAP requires JavaScript form interaction. This will be
        implemented with Playwright if production data is needed.
        For Phase 4, the OCC bulk files provide sufficient well/permit data.
        """
        self.logger.info(
            "OkTAP production data not implemented in Phase 4. Use OCC bulk files for well data, permits, completions."
        )

    # ------------------------------------------------------------------
    # Response routing
    # ------------------------------------------------------------------

    def parse_bulk_file(self, response):
        """Route parsing based on file format in meta."""
        fmt = response.meta["file_format"]
        report_type = response.meta["report_type"]
        dataset_name = response.meta["dataset_name"]

        self.logger.info(f"Parsing {dataset_name} ({fmt}) -- {len(response.body)} bytes")

        if fmt == "csv":
            yield from self._parse_csv(response, report_type)
        elif fmt == "xlsx":
            yield from self._parse_xlsx(response, report_type)
        else:
            self.logger.error(f"Unknown format '{fmt}' for dataset {dataset_name}")

    # ------------------------------------------------------------------
    # CSV parsing
    # ------------------------------------------------------------------

    def _parse_csv(self, response, report_type: str):
        """Parse CSV response body (RBDMS wells, incidents)."""
        text = response.text
        reader = csv.DictReader(io.StringIO(text))

        for row in reader:
            if report_type == "well_data":
                result = self._parse_rbdms_well_row(row, response.url)
                if result is not None:
                    yield result
            elif report_type == "incident_report":
                result = self._parse_incident_row(row, response.url)
                if result is not None:
                    yield result

    def _parse_rbdms_well_row(self, row: dict, source_url: str) -> WellItem | None:
        """Parse a single RBDMS well data CSV row into a WellItem."""
        api_raw = row.get("API_WELL_NUMBER", row.get("api_well_number", "")).strip()
        if not api_raw:
            return None

        spud_date = self._parse_date(row.get("SPUD_DATE", ""))
        completion_date = self._parse_date(row.get("COMPLETION_DATE", ""))

        return WellItem(
            api_number=self.normalize_api_number(api_raw),
            state_code=self.state_code,
            well_name=row.get("WELL_NAME", "").strip(),
            operator_name=row.get("OPERATOR_NAME", "").strip(),
            county=row.get("COUNTY", "").strip(),
            latitude=self._parse_float(row.get("LATITUDE")),
            longitude=self._parse_float(row.get("LONGITUDE")),
            well_status=row.get("WELL_STATUS", "").strip(),
            well_type=row.get("WELL_TYPE", "").strip(),
            spud_date=spud_date,
            completion_date=completion_date,
            total_depth=self._parse_int(row.get("TOTAL_DEPTH")),
            metadata={
                "formation": row.get("FORMATION_NAME", "").strip(),
                "section": row.get("SECTION", "").strip(),
                "township": row.get("TOWNSHIP", "").strip(),
                "range": row.get("RANGE", "").strip(),
                "well_class": row.get("WELL_CLASS", "").strip(),
                "operator_number": row.get("OPERATOR_NUMBER", "").strip(),
                "first_prod_date": row.get("FIRST_PROD_DATE", "").strip(),
                "plug_date": row.get("PLUG_DATE", "").strip(),
                "source_url": source_url,
            },
        )

    def _parse_incident_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse an incident report CSV row into a DocumentItem."""
        api_raw = row.get("API_WELL_NUMBER", "").strip()

        return self.build_document_item(
            source_url=source_url,
            doc_type="incident_report",
            api_number=api_raw if api_raw else None,
            operator_name=row.get("OPERATOR_NAME", "").strip() or None,
            well_name=row.get("WELL_NAME", "").strip() or None,
            raw_metadata={
                "incident_date": row.get("INCIDENT_DATE", "").strip(),
                "incident_type": row.get("INCIDENT_TYPE", "").strip(),
                "county": row.get("COUNTY", "").strip(),
                "description": row.get("DESCRIPTION", "").strip(),
                "resolution": row.get("RESOLUTION", "").strip(),
            },
        )

    # ------------------------------------------------------------------
    # XLSX parsing
    # ------------------------------------------------------------------

    def _parse_xlsx(self, response, report_type: str):
        """Parse XLSX response body using openpyxl."""
        if openpyxl is None:
            self.logger.error("openpyxl not installed -- cannot parse XLSX files")
            return

        workbook = openpyxl.load_workbook(io.BytesIO(response.body), read_only=True, data_only=True)
        sheet = workbook.active

        headers, data_start_row = self._detect_xlsx_headers(sheet)
        if not headers:
            self.logger.error(
                f"No header row found in XLSX for {report_type} ({response.meta.get('dataset_name', 'unknown')})"
            )
            workbook.close()
            return

        for row in sheet.iter_rows(min_row=data_start_row + 1, values_only=True):
            # Skip completely empty rows
            if not any(cell is not None for cell in row):
                continue

            row_dict = dict(zip(headers, row, strict=False))

            if report_type == "well_permit":
                result = self._parse_itd_row(row_dict, response.url)
                if result is not None:
                    yield result
            elif report_type == "completion_report":
                result = self._parse_completion_row(row_dict, response.url)
                if result is not None:
                    yield result
            elif report_type == "operator_list":
                result = self._parse_operator_row(row_dict)
                if result is not None:
                    yield result
            elif report_type == "uic_data":
                result = self._parse_uic_row(row_dict, response.url)
                if result is not None:
                    yield result
            elif report_type == "uic_injection":
                result = self._parse_uic_injection_row(row_dict, response.url)
                if result is not None:
                    yield result
            elif report_type == "well_transfer":
                result = self._parse_transfer_row(row_dict, response.url)
                if result is not None:
                    yield result

        workbook.close()

    def _detect_xlsx_headers(self, sheet) -> tuple[list[str] | None, int]:
        """Find the header row in an XLSX sheet.

        Government Excel files may have title rows, merged cells, or
        blank rows before the actual header. We scan the first 10 rows
        looking for one that contains mostly non-empty string values.

        Returns:
            (headers, data_start_row) or (None, 0) if not found.
        """
        for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
            if row is None:
                continue
            non_none = [cell for cell in row if cell is not None]
            if not non_none:
                continue
            str_count = sum(1 for cell in non_none if isinstance(cell, str))
            # Consider it a header if at least half the non-empty cells are strings
            # and there are at least 2 non-empty cells
            if len(non_none) >= 2 and str_count >= len(non_none) / 2:
                headers = [str(cell).strip().upper() if cell is not None else "" for cell in row]
                return headers, i

        return None, 0

    def _parse_itd_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse Intent to Drill (drilling permit) XLSX row."""
        api_raw = self._get_api_from_row(row)
        if not api_raw:
            return None

        return self.build_document_item(
            source_url=source_url,
            doc_type="well_permit",
            api_number=api_raw,
            operator_name=self._str_val(row.get("OPERATOR_NAME", row.get("OPERATOR", ""))),
            well_name=self._str_val(row.get("WELL_NAME", row.get("WELL", ""))),
            raw_metadata={
                "permit_type": "Intent to Drill",
                "county": self._str_val(row.get("COUNTY", "")),
                "formation": self._str_val(row.get("FORMATION", "")),
                "proposed_depth": self._parse_float(row.get("PROPOSED_DEPTH")),
                "filing_date": self._str_val(row.get("FILING_DATE", "")),
            },
        )

    def _parse_completion_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse well completion XLSX row."""
        api_raw = self._get_api_from_row(row)
        if not api_raw:
            return None

        return self.build_document_item(
            source_url=source_url,
            doc_type="completion_report",
            api_number=api_raw,
            operator_name=self._str_val(row.get("OPERATOR_NAME", row.get("OPERATOR", ""))),
            well_name=self._str_val(row.get("WELL_NAME", row.get("WELL", ""))),
            raw_metadata={
                "completion_date": self._str_val(row.get("COMPLETION_DATE", "")),
                "formation": self._str_val(row.get("FORMATION", "")),
                "total_depth": self._parse_float(row.get("TOTAL_DEPTH")),
                "first_prod_date": self._str_val(row.get("FIRST_PROD_DATE", "")),
                "initial_oil": self._parse_float(row.get("INITIAL_OIL_PROD")),
                "initial_gas": self._parse_float(row.get("INITIAL_GAS_PROD")),
            },
        )

    def _parse_operator_row(self, row: dict) -> dict | None:
        """Parse operator list XLSX row.

        Yields a plain dict (not DocumentItem/WellItem) as operator
        records are reference data rather than well-specific documents.
        """
        operator_name = self._str_val(row.get("OPERATOR_NAME", row.get("OPERATOR", "")))
        if not operator_name:
            return None

        return {
            "type": "operator",
            "state_code": self.state_code,
            "operator_name": operator_name,
            "operator_number": self._str_val(row.get("OPERATOR_NUMBER", row.get("OPERATOR_NUM", ""))),
            "address": self._str_val(row.get("ADDRESS", "")),
            "city": self._str_val(row.get("CITY", "")),
            "state": self._str_val(row.get("STATE", "")),
            "zip_code": self._str_val(row.get("ZIP", row.get("ZIP_CODE", ""))),
        }

    def _parse_uic_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse UIC (Underground Injection Control) well XLSX row."""
        api_raw = self._get_api_from_row(row)
        if not api_raw:
            return None

        return self.build_document_item(
            source_url=source_url,
            doc_type="uic_data",
            api_number=api_raw,
            operator_name=self._str_val(row.get("OPERATOR_NAME", row.get("OPERATOR", ""))),
            well_name=self._str_val(row.get("WELL_NAME", row.get("WELL", ""))),
            raw_metadata={
                "well_class": self._str_val(row.get("WELL_CLASS", "")),
                "county": self._str_val(row.get("COUNTY", "")),
                "permit_number": self._str_val(row.get("PERMIT_NUMBER", row.get("PERMIT_NO", ""))),
                "status": self._str_val(row.get("STATUS", "")),
            },
        )

    def _parse_uic_injection_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse UIC injection volume XLSX row."""
        api_raw = self._get_api_from_row(row)
        if not api_raw:
            return None

        return self.build_document_item(
            source_url=source_url,
            doc_type="uic_injection",
            api_number=api_raw,
            operator_name=self._str_val(row.get("OPERATOR_NAME", row.get("OPERATOR", ""))),
            well_name=self._str_val(row.get("WELL_NAME", row.get("WELL", ""))),
            raw_metadata={
                "injection_volume": self._parse_float(row.get("INJECTION_VOLUME", row.get("VOLUME"))),
                "reporting_period": self._str_val(row.get("REPORTING_PERIOD", row.get("PERIOD", ""))),
                "formation": self._str_val(row.get("FORMATION", "")),
                "pressure": self._parse_float(row.get("PRESSURE")),
            },
        )

    def _parse_transfer_row(self, row: dict, source_url: str) -> DocumentItem | None:
        """Parse well transfer XLSX row."""
        api_raw = self._get_api_from_row(row)
        if not api_raw:
            return None

        return self.build_document_item(
            source_url=source_url,
            doc_type="well_transfer",
            api_number=api_raw,
            raw_metadata={
                "from_operator": self._str_val(row.get("FROM_OPERATOR", row.get("PREV_OPERATOR", ""))),
                "to_operator": self._str_val(row.get("TO_OPERATOR", row.get("NEW_OPERATOR", ""))),
                "transfer_date": self._str_val(row.get("TRANSFER_DATE", "")),
                "well_name": self._str_val(row.get("WELL_NAME", row.get("WELL", ""))),
                "county": self._str_val(row.get("COUNTY", "")),
            },
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_api_from_row(self, row: dict) -> str | None:
        """Extract and validate API number from a row dict.

        Tries common column name variants. Returns None if no
        valid API number is found.
        """
        for key in ("API_WELL_NUMBER", "API", "API_NUMBER", "API_NO"):
            val = row.get(key)
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    return val_str
        return None

    @staticmethod
    def _str_val(value) -> str:
        """Safely convert any XLSX cell value to a stripped string.

        XLSX cells may contain numbers, dates, None, or strings.
        """
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        return str(value).strip()

    @staticmethod
    def _parse_float(value) -> float | None:
        """Parse a value as float, returning None on failure."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_int(value) -> int | None:
        """Parse a value as int, returning None on failure."""
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_date(value) -> date | None:
        """Parse a date string in common formats, returning None on failure."""
        if not value or not isinstance(value, str):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            return None
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y%m%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None
